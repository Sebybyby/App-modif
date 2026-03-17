# -*- coding: utf-8 -*-
"""
Gestion de la base de données des lecteurs Soft Pos (sp_lecteurs.xlsx).
Chaque entrée est un téléphone avec sa position 0 : x, y, z, rX, rY, rZ, topZ.
Les lecteurs matériels (Lane 5000, Move 5000) sont dans LecteurDB.py / lecteurs.xlsx.
"""
import os
import openpyxl

SP_LECTEURS_FILE = "./Parametres/sp_lecteurs.xlsx"
HEADERS = ["Nom", "x", "y", "z", "rX", "rY", "rZ", "topZ"]

# topZ Soft Pos — même valeur que positionTopZLane5000 dans RobotClass.py
_TOP_Z = 0.399821937815115

# Téléphones Soft Pos — coordonnées extraites des méthodes Position0_XXX de RobotClass.py
DEFAULT_SP_LECTEURS = [
    ["iPhone SE",
     -0.4535050711611319,   0.09746050997975206,  0.1628971910211756,
     -2.231798624387687,    2.2085533882615613,   0.012635744971039179,
     _TOP_Z],
    ["iPhone 13",
     -0.4530451975084202,   0.09109465497697321,  0.1623321775801928,
     -2.231812791929215,    2.20836277121875,     0.012548477860910717,
     _TOP_Z],
    ["Samsung A34",
     -0.43754267515397605,  0.05571605275012892,  0.16325538725581454,
     -2.2287899201722428,   2.211582666496985,    0.012739503012607888,
     _TOP_Z],
    ["Samsung S23",
     -0.44212966228298567,  0.0018719860483787305, 0.16401135259168176,
     -2.2120128473028395,   2.2282316325468523,   0.012707245680119298,
     _TOP_Z],
    ["Samsung A15",
     -0.4386549445573203,   0.0774494613480611,   0.16674799989871197,
     -2.2162006686819176,   2.22500596381785,     0.016132300618970388,
     _TOP_Z],
    ["Redmi 13C",
     -0.4342770051859024,   0.07549404718371988,  0.1650135223170027,
      2.2089024541654294,  -2.2273135015530854,   0.0029808527696273787,
     _TOP_Z],
]

_SP_NAMES = {row[0] for row in DEFAULT_SP_LECTEURS}


def _ensure_file():
    """Crée sp_lecteurs.xlsx s'il n'existe pas, puis ajoute les entrées manquantes."""
    os.makedirs("./Parametres", exist_ok=True)
    if not os.path.exists(SP_LECTEURS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SP_Lecteurs"
        ws.append(HEADERS)
        for row in DEFAULT_SP_LECTEURS:
            ws.append(row)
        wb.save(SP_LECTEURS_FILE)
        return

    wb = openpyxl.load_workbook(SP_LECTEURS_FILE)
    ws = wb.active
    existing = {row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    changed = False
    for row in DEFAULT_SP_LECTEURS:
        if row[0] not in existing:
            ws.append(row)
            changed = True
    if changed:
        wb.save(SP_LECTEURS_FILE)


def get_sp_lecteurs():
    """Retourne la liste des noms de lecteurs Soft Pos."""
    _ensure_file()
    wb = openpyxl.load_workbook(SP_LECTEURS_FILE)
    ws = wb.active
    return [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]


def get_sp_lecteur_position(nom):
    """Retourne un dict {x, y, z, rX, rY, rZ, topZ} pour le lecteur Soft Pos donné."""
    _ensure_file()
    wb = openpyxl.load_workbook(SP_LECTEURS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == nom:
            return {
                "x": row[1], "y": row[2], "z": row[3],
                "rX": row[4], "rY": row[5], "rZ": row[6],
                "topZ": row[7],
            }
    return None


def add_sp_lecteur(nom, x, y, z, rX, rY, rZ, topZ):
    """Ajoute ou met à jour un lecteur Soft Pos dans la base de données."""
    _ensure_file()
    wb = openpyxl.load_workbook(SP_LECTEURS_FILE)
    ws = wb.active
    vals = [nom, x, y, z, rX, rY, rZ, topZ]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == nom:
            for i, v in enumerate(vals):
                row[i].value = v
            wb.save(SP_LECTEURS_FILE)
            return
    ws.append(vals)
    wb.save(SP_LECTEURS_FILE)
