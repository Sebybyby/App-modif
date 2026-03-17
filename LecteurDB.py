# -*- coding: utf-8 -*-
"""
Gestion de la base de données des lecteurs Combinatoire (lecteurs.xlsx).
Chaque lecteur a un nom et sa position 0 : x, y, z, rX, rY, rZ, topZ.
Seuls les lecteurs matériels (Lane 5000, Move 5000) sont gérés ici.
Les lecteurs Soft Pos (téléphones) sont dans SP_LecteurDB.py / sp_lecteurs.xlsx.
"""
import os
import openpyxl

LECTEURS_FILE = "./Parametres/lecteurs.xlsx"
HEADERS = ["Nom", "x", "y", "z", "rX", "rY", "rZ", "topZ"]

# topZ constants extraites de RobotClass.py
_TOP_Z_LANE = 0.399821937815115    # positionTopZLane5000
_TOP_Z_MOVE = 0.5025886995214306   # positionTopZmove5000

# Lecteurs Combinatoire uniquement
DEFAULT_LECTEURS = [
    ["Lane 5000",
     -0.4547756256699072,   0.09449841328596138,  0.21046844075096408,
     -2.2162304932532844,   2.2250713469998527,   0.01629001438342314,
     _TOP_Z_LANE],
    ["Move 5000",
     -0.4540141593144729,   0.031580682143320125, 0.30871004420187936,
     -2.2037218160115697,   2.2372876043326744,   0.01616534026994103,
     _TOP_Z_MOVE],
]

_COMBINATOIRE_NAMES = {row[0] for row in DEFAULT_LECTEURS}


def _ensure_file():
    """Crée lecteurs.xlsx s'il n'existe pas.
    Si le fichier existe, ajoute les lecteurs manquants et supprime
    les entrées qui n'appartiennent pas au Combinatoire (ex : téléphones
    qui se trouvaient dans l'ancien fichier unique).
    """
    os.makedirs("./Parametres", exist_ok=True)
    if not os.path.exists(LECTEURS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lecteurs"
        ws.append(HEADERS)
        for row in DEFAULT_LECTEURS:
            ws.append(row)
        wb.save(LECTEURS_FILE)
        return

    wb = openpyxl.load_workbook(LECTEURS_FILE)
    ws = wb.active
    existing = {row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    changed = False

    # Supprimer les lignes qui ne font pas partie du Combinatoire
    rows_to_keep = [ws[1]]  # header
    for row in ws.iter_rows(min_row=2):
        nom = row[0].value
        if nom and nom in _COMBINATOIRE_NAMES:
            rows_to_keep.append(row)
        elif nom:
            changed = True  # ligne supprimée

    if changed:
        # Réécrire la feuille proprement
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        for dest_row, src_row in enumerate(rows_to_keep[1:], start=2):
            for col_idx, cell in enumerate(src_row, start=1):
                ws.cell(row=dest_row, column=col_idx).value = cell.value

    # Ajouter les lecteurs par défaut manquants
    existing_after = {
        ws.cell(row=r, column=1).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    }
    for row in DEFAULT_LECTEURS:
        if row[0] not in existing_after:
            ws.append(row)
            changed = True

    if changed:
        wb.save(LECTEURS_FILE)


def get_lecteurs():
    """Retourne la liste des noms de lecteurs Combinatoire."""
    _ensure_file()
    wb = openpyxl.load_workbook(LECTEURS_FILE)
    ws = wb.active
    return [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]


def get_lecteur_position(nom):
    """Retourne un dict {x, y, z, rX, rY, rZ, topZ} pour le lecteur donné."""
    _ensure_file()
    wb = openpyxl.load_workbook(LECTEURS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == nom:
            return {
                "x": row[1], "y": row[2], "z": row[3],
                "rX": row[4], "rY": row[5], "rZ": row[6],
                "topZ": row[7],
            }
    return None


def add_lecteur(nom, x, y, z, rX, rY, rZ, topZ):
    """Ajoute ou met à jour un lecteur Combinatoire dans la base de données."""
    _ensure_file()
    wb = openpyxl.load_workbook(LECTEURS_FILE)
    ws = wb.active
    vals = [nom, x, y, z, rX, rY, rZ, topZ]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == nom:
            for i, v in enumerate(vals):
                row[i].value = v
            wb.save(LECTEURS_FILE)
            return
    ws.append(vals)
    wb.save(LECTEURS_FILE)
