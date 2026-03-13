import openpyxl
from CompAudioConfig import fftsignallecteur
from InterfaceAfficheClass import InterfaceAffiche
from InterfaceClass import Interface
from PySide6.QtWidgets import (
    QLabel, QPushButton, QLineEdit, QFileDialog, QVBoxLayout, QWidget
)
from PySide6.QtGui import QFont
from PySide6.QtCore import Qt
from openpyxl import Workbook
import os
import gc


class FichierReader(fftsignallecteur, Interface):
    def __init__(self):
        Interface.__init__(self)
        self.NouvellePageLect()
        self.enum = InterfaceAffiche(6)
        self.grid_rowconfigure((0, 1, 2, 3, 4, 5, 6, 7, 8, 9), weight=1)
        self.grid_columnconfigure((0, 1, 2, 3, 4, 5, 6, 7, 8), weight=1)

    def NouvellePageLect(self):
        BG = "#0054A4"
        BTN = ("QPushButton { background:#FFFF00; color:#000000; font-size:13px;"
               " font-weight:bold; padding:12px 40px; border-radius:8px; }"
               "QPushButton:hover    { background:#FFE033; }"
               "QPushButton:pressed  { background:#CCBB00; }")

        # Retour
        self.retourBouton = QPushButton("Retour")
        self.retourBouton.setStyleSheet(
            "QPushButton { background:#FFFF00; color:#000000; font-size:13px;"
            " font-weight:bold; padding:8px 20px; border-radius:6px; }"
            "QPushButton:hover    { background:#FFE033; }"
            "QPushButton:pressed  { background:#CCBB00; }"
        )
        self.retourBouton.clicked.connect(self.RetourMenu)
        self._grid.addWidget(self.retourBouton, 0, 0, Qt.AlignLeft | Qt.AlignTop)

        # Titre centré
        self.ti = QLabel("Détection transaction")
        self.ti.setStyleSheet("color:#FFFFFF; background:transparent; border:none;")
        self.ti.setFont(QFont("Helvetica", 28, QFont.Bold))
        self.ti.setAlignment(Qt.AlignCenter)
        self._grid.addWidget(self.ti, 0, 1, 1, 7, Qt.AlignCenter)

        # Boutons centrés
        self.boutonDetectionReader = QPushButton(
            "Ajout d'un lecteur\nRéglage du Son de Paiement"
        )
        self.boutonDetectionReader.setStyleSheet(BTN)
        self.boutonDetectionReader.clicked.connect(self.Comparaison_totale)
        self._grid.addWidget(self.boutonDetectionReader, 2, 2, 1, 5, Qt.AlignCenter)

        self.entry = QLineEdit()
        self.entry.setPlaceholderText("Nom du lecteur...")
        self.entry.setStyleSheet(
            "QLineEdit { background:#FFFFFF; color:#000000; font-size:13px;"
            " padding:8px; border-radius:6px; }"
        )
        self.entry.setFixedWidth(300)
        self._grid.addWidget(self.entry, 4, 2, 1, 5, Qt.AlignCenter)

        self.boutonStockage = QPushButton("Stockage des paramètres de ce lecteur")
        self.boutonStockage.setStyleSheet(BTN)
        self.boutonStockage.clicked.connect(self.creer_feuille)
        self._grid.addWidget(self.boutonStockage, 6, 2, 1, 5, Qt.AlignCenter)

        self.FileBouton = QPushButton("Lecteur déjà paramétré")
        self.FileBouton.setStyleSheet(BTN)
        self.FileBouton.clicked.connect(self.ouvrir_fichier_excel)
        self._grid.addWidget(self.FileBouton, 8, 2, 1, 5, Qt.AlignCenter)

    def creer_feuille(self):
        dira = self.Verification()
        print(len(dira))
        nb_dir = len(dira)

        classeur = Workbook()
        feuille = classeur.active
        feuille.cell(row=1, column=3, value="Il y a" + str(nb_dir) + "diracs")
        feuille.cell(row=1, column=4, value=nb_dir)
        for i in range(1, len(dira) + 1):
            feuille.cell(row=i, column=1, value="Dirac" + str(i))
            feuille.cell(row=i, column=2, value=dira[i - 1])

        nom_fichier = self.entry.text()
        classeur.close()
        classeur.save("./Parametres/Lecteur/" + nom_fichier + ".xlsx")

        if os.path.exists("./Parametres/Lecteur/" + nom_fichier + ".xlsx"):
            self._show_ok_dialog(f"   Fichier '{nom_fichier}' créé avec succès   ", w=400, h=130)
        else:
            self._show_ok_dialog("   Impossible de créer le fichier Excel   ", w=400, h=130)
            print(nom_fichier)

    def RetourMenu(self):
        self.windowPlace = self.geometry()
        self.destroy()
        gc.collect()
        self.enum = 0

    def ouvrir_fichier_excel(self):
        fichier_excel, _ = QFileDialog.getOpenFileName(
            self, "Ouvrir un fichier Excel", "",
            "Fichiers Excel (*.xlsx *.xls)"
        )
        if fichier_excel:
            classeur = openpyxl.load_workbook(fichier_excel)
            feuille = classeur.active
            cellule_B1 = feuille["B1"].value
            cellule_B2 = feuille["B2"].value
            cellule_B3 = feuille["B3"].value
            Nomb_dir = feuille["D1"].value
            print("Contenu des cellules B1, B2 et B3 :")
            print("B1 :", cellule_B1)
            print("B2 :", cellule_B2)
            print("B3 :", cellule_B3)
            self.cl = Workbook()
            self.sheet = self.cl.active
            self.sheet["B1"] = cellule_B1
            self.sheet["B2"] = cellule_B2
            self.sheet["B3"] = cellule_B3
            self.sheet["D1"] = Nomb_dir
            self.cl.save("./Parametres/Config_Lecteur_Utilise.xlsx")

    def ValueSave(self):
        donneedirac = openpyxl.load_workbook("./Parametres/Config_Lecteur_Utilise.xlsx")
        sh = donneedirac.active
        D1 = sh["B1"].value
        D2 = sh["B2"].value
        D3 = sh["B3"].value
        nb_dir = sh["D1"].value
        return nb_dir, D1, D2, D3
