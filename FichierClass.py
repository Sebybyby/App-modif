# -*- coding: utf-8 -*-
"""
Created on Fri May  7 14:25:08 2021

@author: t0247275
"""
from datetime import datetime
from RobotClass import Robot

import openpyxl
from openpyxl import  load_workbook
from copy import copy


class Fichier():
    instance=None
    def __init__(self):
         self.rob = Robot.Instance()
         
    def creationFichier(self,nomCarte,nomOperateur,nomAntenne):
         carte = nomCarte.text()
         operateur = nomOperateur.text()
         antenne = nomAntenne.text()
         self.titre = "./Projet/"+carte+".txt"
         self.Fichierxlsx=load_workbook('./Parametres/MODELE.xlsx')
         self.titrexlsx = "./Projet/"+carte+".xlsx"
         self.Fichierxlsx.save(self.titrexlsx)
         self.worksheet = self.Fichierxlsx.active

         self.worksheet.cell(1,2,carte)
         self.worksheet.cell(2,2,antenne)
         self.worksheet.cell(3,2,operateur)
         print(self.titre)
         self.fichier = open(self.titre,"w")
         nom = "Le nom de l'opérateur est : " + operateur + "\n"
         fichierAntenne = "le nom de l'antenne est : " +  antenne + "\n"
         fichierProjet = "le nom du projet est : " + carte + "\n"
         self.fichier.write (nom)
         self.fichier.write (fichierAntenne)
         self.fichier.write (fichierProjet)
         if (self.rob.offsetX != 0 or self.rob.offsetY != 0 or self.rob.offsetZ != 0): 
             self.fichier.write ("\nATTENTION offset appliqué(s) à toutes les positions : offsetX = "+str(self.rob.offsetX)+"m ; offsetY = "+str(self.rob.offsetY)+"m ; offsetZ = "+str(self.rob.offsetZ)+"m\n")
         self.fichier.write ("\n")
         self.fichier.close()
         self.Fichierxlsx.save(self.titrexlsx)
         

    def GroupeEcriture(self, i):
        self.fichier = open(self.titre,"a")
        if (i==0):
            self.fichier.write("\n-------- Groupe A --------\n")
        elif(i==11):
            self.fichier.write("\n-------- Groupe B --------\n")
        elif(i==24):
            self.fichier.write("\n-------- Groupe C --------\n")
        elif(i==37):
            self.fichier.write("\n-------- Groupe D --------\n")        
        elif(i==50):
             self.fichier.write("\n-------- Groupe E --------\n")       
        self.fichier.close()

        
    def Lane5000(self):
         self.reader = 1
         self.fichier = open(self.titre,"a")
         self.fichier.write("\n-------- LANE 5000 --------\n")     
         self.fichier.close()
    def Move5000(self):
        self.reader = 0
        self.fichier = open(self.titre,"a")
        self.fichier.write("\n-------- MOVE 5000 --------\n")
        self.fichier.close()

    def EcritureLecteur(self, nom):
        self.fichier = open(self.titre,"a")
        self.fichier.write(f"\n-------- {nom} --------\n")
        self.fichier.close()
    def Manuel(self):
        self.fichier = open(self.titre,"a")
        self.fichier.write("\n-------- Mode : Manuel --------\n")  
        self.fichier.close()
    def Auto(self):
        self.fichier = open(self.titre,"a")
        self.fichier.write("\n-------- Mode : Automatique --------\n")  
        self.fichier.close()
    def TransactionPass(self,i,cardloop):
        self.fichier = open(self.titre,"a")
        today = datetime.now()
        ecriture = str(today) +" " +str(self.rob.tab_Coordonnee[i][0]) +", coordonne z : "+ str(self.rob.tab_Coordonnee[i][1]) + ", coordonnee r : "+ str(self.rob.tab_Coordonnee[i][2])+ ", coordonnee phi : "+ str(self.rob.tab_Coordonnee[i][3])+ " : PASS\n"    
        self.fichier.write(ecriture)
        self.fichier.close()
        self.choixsheet(cardloop)
        self.worksheet.cell(i+6,7,1)
        self.Fichierxlsx.save(self.titrexlsx)
    
    def TransactionFail(self,i,cardloop):
        self.fichier = open(self.titre,"a")
        today = datetime.now()
        ecriture = str(today) +" " +str(self.rob.tab_Coordonnee[i][0]) +", coordonne z : "+ str(self.rob.tab_Coordonnee[i][1]) + ", coordonnee r : "+ str(self.rob.tab_Coordonnee[i][2])+ ", coordonnee phi : "+ str(self.rob.tab_Coordonnee[i][3])+ " : FAIL\n"    
        self.fichier.write(ecriture)
        self.fichier.close()
        self.choixsheet(cardloop)
        self.worksheet.cell(i+6,7,0)
        self.Fichierxlsx.save(self.titrexlsx)
                
    def ModeAutomatique(self):
        self.fichier = open(self.titre,"a")
        ecriture = "--------Mode Automatique--------" 
        self.fichier.write(ecriture)
        self.fichier.close()
     
    def ModeManuel(self):
        self.fichier = open(self.titre,"a")
        ecriture = "--------Mode Manuel--------" 
        self.fichier.write(ecriture)
        self.fichier.close()
    
    def choixsheet(self,cardloop):
        self.Fichierxlsx.active = self.Fichierxlsx.worksheets[cardloop]
        self.worksheet = self.Fichierxlsx.active
        self.Fichierxlsx.save(self.titrexlsx)
        
    @staticmethod
    def Instance():
        if (Fichier.instance==None):
            Fichier.instance = Fichier()
        return Fichier.instance